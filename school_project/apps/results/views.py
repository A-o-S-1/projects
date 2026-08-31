import csv
import io
from datetime import timedelta

import openpyxl
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views import View
from django.views.generic import FormView

from . import services
from .forms import MasterSheetForm, ResultLookupForm, ScoreCSVUploadForm, WorkbookUploadForm
from .models import (
    AcademicSession, ClassRoom, ResultCheckLog, ResultEntry, PsychomotorRating,
    ScratchCard, SessionResult, Student, Term, TermResult,
)
from apps.academics.models import Subject

# How many failed attempts from one IP within the window before we block
# further tries. Deliberately generous enough not to lock out a parent who
# fat-fingers a PIN twice, strict enough to make brute-forcing impractical.
RATE_LIMIT_MAX_ATTEMPTS = 5
RATE_LIMIT_WINDOW_MINUTES = 15

# How long a successful lookup stays viewable before the parent has to
# check again with the (now-used-up) card. Short enough that a result
# left open on a shared/library computer doesn't stay exposed for long.
SESSION_UNLOCK_MINUTES = 15

SESSION_KEY = "unlocked_term_result_id"
SESSION_KEY_EXPIRES = "unlocked_term_result_expires"


def _client_ip(request):
    """
    Best-effort real client IP behind a reverse proxy. Checked against
    X-Forwarded-For first (set by Nginx/the hosting platform in production),
    falling back to REMOTE_ADDR for local development.
    """
    forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR", "0.0.0.0")


def _is_rate_limited(ip_address):
    window_start = timezone.now() - timedelta(minutes=RATE_LIMIT_WINDOW_MINUTES)
    recent_failed = ResultCheckLog.objects.filter(
        ip_address=ip_address, was_successful=False, attempted_at__gte=window_start
    ).count()
    return recent_failed >= RATE_LIMIT_MAX_ATTEMPTS


class ResultLookupView(FormView):
    """
    Public result lookup: Admission Number + Term + Scratch Card PIN + Serial.

    Security properties (see also the module-level constants above):
    - Every attempt is logged (ResultCheckLog), success or failure.
    - Failures never reveal WHICH field was wrong — same generic message
      whether the admission number doesn't exist, the PIN is wrong, the
      card was already used, or the result simply isn't published yet.
      Distinguishing these would let an attacker enumerate valid admission
      numbers or valid-but-unused PINs one field at a time.
    - IP-based rate limiting via the same log table — no new dependency.
    - On success, the card is immediately marked used (single-use) and the
      result becomes viewable only through a time-limited session flag,
      not a bookmarkable/guessable URL.
    """
    template_name = "results/lookup.html"
    form_class = ResultLookupForm

    def form_valid(self, form):
        ip_address = _client_ip(self.request)
        admission_number = form.cleaned_data["admission_number"]

        if _is_rate_limited(ip_address):
            form.add_error(None, "Too many attempts. Please wait 15 minutes before trying again.")
            return self.render_to_response(self.get_context_data(form=form))

        term = Term.objects.filter(pk=form.cleaned_data["term"]).first()
        student = Student.objects.filter(admission_number__iexact=admission_number, is_active=True).first()
        card = ScratchCard.objects.filter(
            serial_number__iexact=form.cleaned_data["serial_number"],
            pin=form.cleaned_data["pin"],
            is_used=False,
        ).first()

        term_result = None
        if student and term:
            term_result = TermResult.objects.filter(student=student, term=term).first()

        success = bool(student and term and card and term_result and term_result.is_publicly_visible)

        ResultCheckLog.objects.create(
            admission_number_attempted=admission_number, ip_address=ip_address, was_successful=success
        )

        if not success:
            form.add_error(
                None,
                "We couldn't find a result matching those details. Please double-check your admission "
                "number, term, PIN, and serial number, or confirm the result has been published.",
            )
            return self.render_to_response(self.get_context_data(form=form))

        card.mark_used(student)
        self.request.session[SESSION_KEY] = term_result.id
        self.request.session[SESSION_KEY_EXPIRES] = (
            timezone.now() + timedelta(minutes=SESSION_UNLOCK_MINUTES)
        ).isoformat()
        return redirect(reverse("results:view_result"))


class ResultDetailView(View):
    """
    Displays the unlocked result — reachable only via a valid session flag
    set by ResultLookupView, never by a direct/guessable URL. Session flag
    is time-limited and single-use per successful card check.
    """
    template_name = "results/result_detail.html"

    def get(self, request):
        term_result_id = request.session.get(SESSION_KEY)
        expires_at = request.session.get(SESSION_KEY_EXPIRES)

        if not term_result_id or not expires_at or timezone.now().isoformat() > expires_at:
            request.session.pop(SESSION_KEY, None)
            request.session.pop(SESSION_KEY_EXPIRES, None)
            return redirect(reverse("results:check"))

        term_result = TermResult.objects.filter(pk=term_result_id, is_published=True, is_blocked=False).first()
        if not term_result:
            return redirect(reverse("results:check"))

        result_entries = ResultEntry.objects.filter(
            student=term_result.student, term=term_result.term
        ).select_related("subject")
        psychomotor_ratings = PsychomotorRating.objects.filter(
            student=term_result.student, term=term_result.term
        ).select_related("skill")

        return render(request, self.template_name, {
            "term_result": term_result,
            "student": term_result.student,
            "result_entries": result_entries,
            "psychomotor_ratings": psychomotor_ratings,
        })


class ResultLogoutView(View):
    """Lets the parent explicitly clear the unlocked result before the
    session naturally expires — e.g. on a shared computer."""

    def get(self, request):
        request.session.pop(SESSION_KEY, None)
        request.session.pop(SESSION_KEY_EXPIRES, None)
        return redirect(reverse("results:check"))


# ==============================================================================
# Staff-only tools: CSV score upload, position recalculation, master sheets
# ==============================================================================
@method_decorator(staff_member_required, name="dispatch")
class ScoreUploadView(View):
    """
    Bulk score upload for a whole class/term at once from a CSV file,
    rather than entering each ResultEntry one at a time in the admin.

    Gated with @staff_member_required (Django's built-in admin-login check)
    rather than building separate authentication — this reuses the same
    login every other admin action already requires.
    """
    template_name = "results/staff/upload_scores.html"

    def get(self, request):
        return render(request, self.template_name, {"form": ScoreCSVUploadForm()})

    def post(self, request):
        form = ScoreCSVUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, {"form": form})

        term = Term.objects.get(pk=form.cleaned_data["term"])
        csv_file = form.cleaned_data["csv_file"]

        decoded = csv_file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))

        created, updated, errors = 0, 0, []
        touched_students = set()

        required_columns = {"admission_number", "subject", "ca_score", "exam_score"}
        if not required_columns.issubset(set(reader.fieldnames or [])):
            errors.append(
                f"CSV header must include: {', '.join(sorted(required_columns))}. "
                f"Found: {', '.join(reader.fieldnames or []) or '(empty file)'}"
            )
            return render(request, self.template_name, {"form": form, "errors": errors})

        for row_number, row in enumerate(reader, start=2):  # row 1 is the header
            admission_number = (row.get("admission_number") or "").strip()
            subject_name = (row.get("subject") or "").strip()

            student = Student.objects.filter(admission_number__iexact=admission_number).first()
            if not student:
                errors.append(f"Row {row_number}: no student with admission number '{admission_number}'.")
                continue

            # Subject names like "Mathematics" exist at BOTH junior and
            # senior level — resolve using the student's own class level so
            # a JSS1 student's "Mathematics" never gets attached to the
            # senior-level Subject record (or vice versa).
            subject_qs = Subject.objects.filter(name__iexact=subject_name)
            if student.current_class:
                subject = subject_qs.filter(level=student.current_class.subject_level).first()
                if not subject:
                    # Fall back to any match (e.g. the subject only exists
                    # at one level) rather than failing outright.
                    subject = subject_qs.first()
            else:
                subject = subject_qs.first()

            if not subject:
                errors.append(f"Row {row_number}: no subject named '{subject_name}'.")
                continue

            try:
                ca_score = float(row.get("ca_score") or 0)
                exam_score = float(row.get("exam_score") or 0)
            except ValueError:
                errors.append(f"Row {row_number}: CA/Exam score must be a number.")
                continue

            if ca_score > 30 or exam_score > 70 or ca_score < 0 or exam_score < 0:
                errors.append(
                    f"Row {row_number}: scores out of range (CA max 30, Exam max 70) for {admission_number}."
                )
                continue

            entry, was_created = ResultEntry.objects.update_or_create(
                student=student, term=term, subject=subject,
                defaults={"ca_score": ca_score, "exam_score": exam_score},
            )
            created += 1 if was_created else 0
            updated += 0 if was_created else 1
            touched_students.add(student.id)

        for student_id in touched_students:
            services.recalculate_term_result_totals(Student.objects.get(pk=student_id), term)
        entries_updated, term_results_updated = services.recalculate_positions_for_term(term)

        return render(request, self.template_name, {
            "form": ScoreCSVUploadForm(),
            "summary": {
                "created": created,
                "updated": updated,
                "students_touched": len(touched_students),
                "entries_ranked": entries_updated,
                "term_results_ranked": term_results_updated,
            },
            "errors": errors,
        })


@method_decorator(staff_member_required, name="dispatch")
class MasterSheetView(View):
    """
    Printable class-wide broadsheet: every active student in a class, every
    subject as a column, for one term — for teachers/admin, not parents.
    Distinct from the individual result sheet parents see via the public
    lookup (results/result_detail.html).
    """
    template_name = "results/staff/master_sheet.html"

    def get(self, request):
        term_id = request.GET.get("term")
        classroom_id = request.GET.get("classroom")

        if not term_id or not classroom_id:
            return render(request, self.template_name, {"form": MasterSheetForm()})

        term = Term.objects.get(pk=term_id)
        classroom = ClassRoom.objects.get(pk=classroom_id)
        students = list(classroom.students.filter(is_active=True).order_by("last_name", "first_name"))

        subjects = list(
            Subject.objects.filter(
                result_entries__student__in=students, result_entries__term=term
            ).distinct().order_by("name")
        )

        entries_by_student = {}
        for student in students:
            entries = {
                e.subject_id: e
                for e in ResultEntry.objects.filter(student=student, term=term)
            }
            term_result = TermResult.objects.filter(student=student, term=term).first()
            entries_by_student[student.id] = {"entries": entries, "term_result": term_result}

        return render(request, self.template_name, {
            "form": MasterSheetForm(initial={"term": term_id, "classroom": classroom_id}),
            "term": term,
            "classroom": classroom,
            "students": students,
            "subjects": subjects,
            "entries_by_student": entries_by_student,
        })


TERM_SHEET_NAMES = {"1ST TERM": "first", "2ND TERM": "second", "3RD TERM": "third"}
SUBJECT_BLOCK_START_COL = 2  # 0-indexed: col 0=admission_no, col 1=student name, col 2 onward=subjects
SUBJECT_BLOCK_WIDTH = 4      # CA, Exam, Total, Grade


@method_decorator(staff_member_required, name="dispatch")
class WorkbookUploadView(View):
    """
    Uploads the school's real result workbook (REGISTER + 1ST/2ND/3RD TERM +
    BROADSHEET sheets) for one class/arm at a time. In one pass:
      1. REGISTER  -> creates any Student who doesn't already exist
      2. each TERM sheet present -> creates/updates ResultEntry rows,
         skipping subjects with no score entered (a school uploading
         mid-term won't have every subject filled in yet)
      3. BROADSHEET -> creates/updates the SessionResult (annual summary)

    Only step 1 requires the REGISTER sheet; any of the other sheets can
    be absent (e.g. uploading First Term only, before Second/Third exist).
    """
    template_name = "results/staff/upload_workbook.html"

    def get(self, request):
        return render(request, self.template_name, {"form": WorkbookUploadForm()})

    def post(self, request):
        form = WorkbookUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, {"form": form})

        session = AcademicSession.objects.get(pk=form.cleaned_data["session"])
        workbook_file = form.cleaned_data["workbook_file"]

        try:
            wb = openpyxl.load_workbook(workbook_file, data_only=True)
        except Exception as exc:
            return render(request, self.template_name, {
                "form": form, "errors": [f"Could not read this file as an Excel workbook: {exc}"],
            })

        errors = []
        summary = {"students_created": 0, "students_updated": 0, "scores_written": 0, "terms_processed": []}

        classroom, class_errors = self._process_register(wb, summary)
        errors.extend(class_errors)

        if classroom:
            for sheet_name, term_key in TERM_SHEET_NAMES.items():
                if sheet_name in wb.sheetnames:
                    term_errors = self._process_term_sheet(wb[sheet_name], term_key, session, classroom, summary)
                    errors.extend(term_errors)
                    if not term_errors:
                        summary["terms_processed"].append(sheet_name)

            if "BROADSHEET" in wb.sheetnames:
                broadsheet_errors = self._process_broadsheet(wb["BROADSHEET"], session, summary)
                errors.extend(broadsheet_errors)

        return render(request, self.template_name, {
            "form": WorkbookUploadForm(),
            "summary": summary if classroom else None,
            "classroom": classroom,
            "errors": errors,
        })

    # ------------------------------------------------------------------
    def _process_register(self, wb, summary):
        """Parses the REGISTER sheet: identifies the class/arm, creates any
        missing Student records. Returns (classroom_or_None, errors)."""
        if "REGISTER" not in wb.sheetnames:
            return None, ["This workbook has no REGISTER sheet — cannot identify the class or students."]

        ws = wb["REGISTER"]
        errors = []

        # Class code is a labeled cell, e.g. "Class Code:" | "J1A" — search
        # the first few rows for it rather than assuming a fixed cell,
        # since the register's header layout could shift slightly.
        class_code = None
        for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            for i, cell in enumerate(row):
                if cell and str(cell).strip().lower().startswith("class code"):
                    if i + 1 < len(row) and row[i + 1]:
                        class_code = str(row[i + 1]).strip()
        if not class_code:
            return None, ["Could not find 'Class Code:' in the REGISTER sheet."]

        classroom_name, level = services.classroom_code_to_name(class_code)
        if not classroom_name:
            return None, [f"Class code '{class_code}' doesn't match the expected pattern (e.g. J1A, S2B)."]

        classroom, _ = ClassRoom.objects.get_or_create(
            name=classroom_name,
            defaults={"level": classroom_name[:-1].lower(), "arm": classroom_name[-1]},
        )

        # Find the header row (contains "Admission No." literally), then read
        # every row below it until a blank admission number ends the list.
        header_row_index = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            if any(cell and "admission" in str(cell).lower() for cell in row):
                header_row_index = row_idx
                break
        if not header_row_index:
            return classroom, ["Could not find the 'Admission No.' header row in REGISTER."]

        for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
            admission_number = (str(row[1]).strip() if row[1] else "")
            if not admission_number:
                continue
            surname = (str(row[2]).strip() if row[2] else "")
            first_name = (str(row[3]).strip() if row[3] else "")
            if not surname or not first_name:
                errors.append(f"Register row for '{admission_number}' is missing a surname or first name — skipped.")
                continue

            student, was_created = Student.objects.get_or_create(
                admission_number=admission_number,
                defaults={"first_name": first_name, "last_name": surname, "current_class": classroom},
            )
            if was_created:
                summary["students_created"] += 1
            else:
                # Keep an existing student's class in sync with this register.
                if student.current_class_id != classroom.id:
                    student.current_class = classroom
                    student.save(update_fields=["current_class"])
                summary["students_updated"] += 1

        return classroom, errors

    @staticmethod
    def _level_code(classroom_name):
        return "junior" if classroom_name.startswith("JSS") else "senior"

    # ------------------------------------------------------------------
    def _process_term_sheet(self, ws, term_key, session, classroom, summary):
        """Parses one term sheet (e.g. '1ST TERM'): writes ResultEntry rows
        for every subject column that has an actual score, skipping blanks."""
        errors = []
        term, _ = Term.objects.get_or_create(session=session, name=term_key)
        level = self._level_code(classroom.name)

        # Row 3 holds subject names (merged across each 4-column block),
        # row 4 holds CA/Exam/Total/Grade sub-headers, data starts row 5.
        header_row = [cell.value for cell in ws[3]]
        subject_columns = []  # (subject, ca_col_index)
        col = SUBJECT_BLOCK_START_COL
        while col < len(header_row) and header_row[col]:
            header_text = header_row[col]
            if header_text in ("Overall Total", "Subject Evaluated", "Subject Offered", "Average", "Position"):
                break
            subject = services.resolve_subject(header_text, level)
            if not subject:
                errors.append(f"{term.get_name_display()}: unrecognized subject column '{header_text}' — skipped.")
            else:
                subject_columns.append((subject, col))
            col += SUBJECT_BLOCK_WIDTH

        touched_students = set()
        for row in ws.iter_rows(min_row=5, values_only=True):
            admission_number = (str(row[0]).strip() if row[0] else "")
            if not admission_number:
                continue
            student = Student.objects.filter(admission_number__iexact=admission_number).first()
            if not student:
                errors.append(f"{term.get_name_display()}: no student '{admission_number}' — run REGISTER import first.")
                continue

            for subject, ca_col in subject_columns:
                ca_score, exam_score = row[ca_col], row[ca_col + 1]
                if ca_score is None and exam_score is None:
                    continue  # subject not yet evaluated for this student — leave it out, not zero
                ResultEntry.objects.update_or_create(
                    student=student, term=term, subject=subject,
                    defaults={"ca_score": ca_score or 0, "exam_score": exam_score or 0},
                )
                summary["scores_written"] += 1
            touched_students.add(student.id)

        for student_id in touched_students:
            services.recalculate_term_result_totals(Student.objects.get(pk=student_id), term)
        services.recalculate_positions_for_term(term)

        return errors

    # ------------------------------------------------------------------
    def _process_broadsheet(self, ws, session, summary):
        """Parses the BROADSHEET sheet: one row per student with the
        finalized annual totals and the staff's promotion decision."""
        errors = []
        header_row_index = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
            if any(cell and "admission" in str(cell).lower() for cell in row):
                header_row_index = row_idx
                break
        if not header_row_index:
            errors.append("Could not find the header row in BROADSHEET — session summary not imported.")
            return errors

        promotion_map = {
            "PROMOTED": "promoted", "PROMOTED_ON_TRIAL": "promoted_on_trial", "REPEAT": "repeat",
        }

        for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
            admission_number = (str(row[1]).strip() if row[1] else "")
            if not admission_number:
                continue
            student = Student.objects.filter(admission_number__iexact=admission_number).first()
            if not student:
                errors.append(f"BROADSHEET: no student '{admission_number}'.")
                continue

            raw_status = (str(row[9]).strip().upper() if len(row) > 9 and row[9] else "")
            SessionResult.objects.update_or_create(
                student=student, session=session,
                defaults={
                    "first_term_total": row[3], "second_term_total": row[4], "third_term_total": row[5],
                    "cumulative_total": row[6], "session_average": row[7],
                    "overall_position": str(row[8]) if row[8] is not None else "",
                    "promotion_status": promotion_map.get(raw_status, ""),
                },
            )
        return errors


# ==============================================================================
# Staff-only printing: individual results and whole-class bulk printing,
# without a scratch card — for form teachers/admin who already have
# legitimate access to every student's record, not just their own child's.
# ==============================================================================
@method_decorator(staff_member_required, name="dispatch")
class StaffResultPrintView(View):
    """One student's result sheet, reusing the same partial the public
    lookup flow renders — staff just skip the PIN/serial check entirely."""
    template_name = "results/staff/result_print.html"

    def get(self, request, student_id, term_id):
        student = Student.objects.filter(pk=student_id).first()
        term = Term.objects.filter(pk=term_id).first()
        if not student or not term:
            return render(request, self.template_name, {"not_found": True})

        term_result = TermResult.objects.filter(student=student, term=term).first()
        result_entries = ResultEntry.objects.filter(student=student, term=term).select_related("subject")
        psychomotor_ratings = PsychomotorRating.objects.filter(
            student=student, term=term
        ).select_related("skill")

        return render(request, self.template_name, {
            "term_result": term_result,
            "student": student,
            "result_entries": result_entries,
            "psychomotor_ratings": psychomotor_ratings,
        })


@method_decorator(staff_member_required, name="dispatch")
class StaffClassResultsPrintView(View):
    """
    Every active student's result sheet in a class, for one term, stacked
    onto one printable page with a page-break between each — this is what
    a form teacher prints once to have a paper copy of their whole class.
    """
    template_name = "results/staff/class_results_print.html"

    def get(self, request):
        term_id = request.GET.get("term")
        classroom_id = request.GET.get("classroom")

        if not term_id or not classroom_id:
            return render(request, self.template_name, {"form": MasterSheetForm()})

        term = Term.objects.get(pk=term_id)
        classroom = ClassRoom.objects.get(pk=classroom_id)
        students = list(classroom.students.filter(is_active=True).order_by("last_name", "first_name"))

        sheets = []
        for student in students:
            term_result = TermResult.objects.filter(student=student, term=term).first()
            result_entries = ResultEntry.objects.filter(student=student, term=term).select_related("subject")
            psychomotor_ratings = PsychomotorRating.objects.filter(
                student=student, term=term
            ).select_related("skill")
            sheets.append({
                "student": student,
                "term_result": term_result,
                "result_entries": result_entries,
                "psychomotor_ratings": psychomotor_ratings,
            })

        return render(request, self.template_name, {
            "form": MasterSheetForm(initial={"term": term_id, "classroom": classroom_id}),
            "term": term,
            "classroom": classroom,
            "sheets": sheets,
        })
